import PySimpleGUI as sg
import openpyxl
from openpyxl import Workbook
from tkinter import filedialog
import os

# ファイルと出力先の選択ダイアログ
def get_file_and_output():
    layout = [
        [sg.Text('Excelファイルを選択してください:'), sg.Input(key='-IN-', disabled=True), sg.FileBrowse(key='-FILE_BROWSE-')],
        [sg.Text('出力ファイル名を指定してください:'), sg.Input(key='-OUT_NAME-')],
        [sg.Text('出力ファイルの保存先を選択してください:'), sg.Input(key='-OUT_DIR-', disabled=True), sg.FolderBrowse(key='-FOLDER_BROWSE-')],
        [sg.Button('次へ'), sg.Button('キャンセル')]
    ]
    window = sg.Window('Excel操作', layout)
    while True:
        event, values = window.read()
        if event == '次へ' and values['-FILE_BROWSE-'] and values['-OUT_NAME-'] and values['-FOLDER_BROWSE-']:
            break
        elif event in (sg.WINDOW_CLOSED, 'キャンセル'):
            window.close()
            return None, None, None
    window.close()
    return values['-FILE_BROWSE-'], values['-OUT_NAME-'], values['-FOLDER_BROWSE-']

# シート選択ダイアログを作成する関数
def select_sheet(wb):
    sheet_names = wb.sheetnames  # シート名の一覧を取得
    layout = [
        [sg.Text('シートを選択してください:')],
        [sg.Combo(sheet_names, key='-SHEET-')],
        [sg.Button('次へ'), sg.Button('キャンセル')]
    ]
    window = sg.Window('シート選択', layout)
    while True:
        event, values = window.read()
        if event == '次へ' and values['-SHEET-']:
            break
        elif event in (sg.WINDOW_CLOSED, 'キャンセル'):
            window.close()
            return None
    window.close()
    return values['-SHEET-']

# 数値を削除して新しいシートにコピーする関数
def copy_without_numbers(ws, new_ws):
    for row in ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column)
            # セルが数値でなければコピー
            if not isinstance(cell.value, (int, float)):
                new_cell.value = cell.value

def main():
    try:
        excel_path, output_name, output_dir = get_file_and_output()
        if not excel_path or not output_name or not output_dir:
            sg.popup_error('ファイルまたは出力先が選択されていません')
            return

        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        sg.popup_error(f'Excelファイルを開く際にエラーが発生しました: {e}')
        return

    try:
        selected_sheet = select_sheet(wb)
        if not selected_sheet:
            sg.popup_error('シートが選択されていません')
            return

        ws = wb[selected_sheet]
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = selected_sheet + "_copy"

        copy_without_numbers(ws, new_ws)
    except Exception as e:
        sg.popup_error(f'シートの操作中にエラーが発生しました: {e}')
        return

    try:
        output_path = os.path.join(output_dir, output_name)
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'
        new_wb.save(output_path)
        sg.popup('成功', '新しいExcelファイルが作成されました')
    except Exception as e:
        sg.popup_error(f'ファイルの保存中にエラーが発生しました: {e}')

if __name__ == "__main__":
    main()